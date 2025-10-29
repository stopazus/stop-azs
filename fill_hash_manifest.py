#!/usr/bin/env python3
"""Automate remediation of the SHA-256 manifest worksheet.

This script locates the manifest workbook, iterates over each evidence row, and
calculates SHA-256 digests for files that exist on disk. When a digest is
generated, the `sha_256_checksum` and `generated_timestamp_(utc)` columns are
populated automatically. Existing values are left intact unless the
``--force`` flag is provided.

Example usage (run from the directory that contains the manifest and files):

    python fill_hash_manifest.py

You can also specify an alternate manifest path or evidence directory:

    python fill_hash_manifest.py --manifest /path/to/SHA256_Manifest_7065f609.xlsx \
        --evidence-dir /path/to/evidence
"""

from __future__ import annotations

import argparse
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, Tuple

import openpyxl

MANIFEST_FILENAME = "SHA256_Manifest_7065f609.xlsx"

EXPECTED_HEADERS = {
    "file_name": "file_name",
    "sha_256_checksum": "sha_256_checksum",
    "generated_timestamp_(utc)": "generated_timestamp_(utc)",
    "verification_method": "verification_method",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--manifest",
        type=Path,
        default=Path(MANIFEST_FILENAME),
        help="Path to the SHA-256 manifest workbook.",
    )
    parser.add_argument(
        "--evidence-dir",
        type=Path,
        help=(
            "Directory that contains the evidence files listed in the manifest. "
            "Defaults to the manifest parent directory."
        ),
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Recalculate hashes even when a checksum already exists.",
    )
    parser.add_argument(
        "--verification-method",
        help=(
            "Value to populate in the verification_method column when it is "
            "currently empty."
        ),
    )
    return parser.parse_args()


def normalize_header(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def column_map(headers: Iterable[Tuple[int, object]]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for column_index, header_value in headers:
        normalized = normalize_header(header_value)
        if normalized:
            mapping[normalized] = column_index
    return mapping


def resolve_manifest_path(manifest_path: Path) -> Path:
    if manifest_path.exists():
        return manifest_path
    raise FileNotFoundError(f"Manifest not found: {manifest_path}")


def compute_sha256(file_path: Path) -> str:
    digest = hashlib.sha256()
    with file_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def main() -> int:
    args = parse_args()
    manifest_path = resolve_manifest_path(args.manifest)
    evidence_dir = args.evidence_dir or manifest_path.parent

    workbook = openpyxl.load_workbook(manifest_path)
    worksheet = workbook.active

    headers = column_map(((cell.col_idx, cell.value) for cell in worksheet[1]))

    missing_columns = [name for name in EXPECTED_HEADERS if EXPECTED_HEADERS[name] not in headers]
    if missing_columns:
        missing = ", ".join(missing_columns)
        raise ValueError(
            f"Manifest is missing required column headers: {missing}. Ensure row 1 contains the expected labels."
        )

    file_col = headers[EXPECTED_HEADERS["file_name"]]
    sha_col = headers[EXPECTED_HEADERS["sha_256_checksum"]]
    timestamp_col = headers[EXPECTED_HEADERS["generated_timestamp_(utc)"]]
    verification_col = headers.get(EXPECTED_HEADERS["verification_method"])

    updated = 0
    skipped_existing = 0
    missing_files = []

    for row in range(2, worksheet.max_row + 1):
        file_cell = worksheet.cell(row=row, column=file_col)
        file_name_raw = file_cell.value
        file_name = str(file_name_raw).strip() if file_name_raw else ""

        if not file_name:
            continue

        file_path = (evidence_dir / file_name).expanduser()
        if not file_path.exists():
            missing_files.append(file_name)
            continue

        sha_cell = worksheet.cell(row=row, column=sha_col)
        current_sha = str(sha_cell.value).strip() if sha_cell.value else ""

        if current_sha and not args.force:
            skipped_existing += 1
            continue

        checksum = compute_sha256(file_path)
        sha_cell.value = checksum
        worksheet.cell(row=row, column=timestamp_col).value = utc_timestamp()

        if args.verification_method and verification_col is not None:
            verification_cell = worksheet.cell(row=row, column=verification_col)
            if not str(verification_cell.value or "").strip():
                verification_cell.value = args.verification_method

        updated += 1

    workbook.save(manifest_path)

    print(f"✅ Updated {updated} record(s) in {manifest_path.name}")
    if skipped_existing:
        print(f"ℹ️  Skipped {skipped_existing} row(s) because a checksum already existed. Use --force to overwrite.")
    if missing_files:
        print("⚠️  Files listed in the manifest but missing on disk:")
        for missing in missing_files:
            print(f"   - {missing}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
